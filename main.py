import openpyxl
from tkinter import filedialog
from tkinter import Tk
import wordlenght
import wordfunc

root = Tk()
root.withdraw()

file_path = filedialog.askopenfilename(
    title="select file",
    filetypes=[("Excel", "*.xlsx"), ("Word", "*.docx"), ("all files", "*.*")]
)
excelheight=0
excelwidth=0
truedatah=0
truedatahw=0

checkclass=wordlenght.check()

checkifemplty=0
checkifempltywidth=0
checkwidth=0
checkheight=0
forwidth=10000
forheight=10000

maxlenght=0

if file_path:
    print("file:", file_path)
    sstr = file_path.rfind('/')
    sub = file_path[sstr + 1:-1]

    excel = openpyxl.load_workbook(file_path)

    print("what we have:")
    for sheet_name in excel.sheetnames:
        print(sheet_name)

    try:
        name=excel.sheetnames
        for data in name:
            data = excel[data]
            for i in range(forwidth):

                for j in range(forheight):
                    value = data.cell(row=i + 1, column=j + 1).value
                    if value is not None:

                        checklenght=checkclass.lenght(str(value))

                        if checklenght>maxlenght:
                            maxlenght=checklenght
                        if excelwidth < j:
                            excelwidth = j+1
                        if excelheight < i:
                            excelheight = i+1
                        checkifemplty = 0
                        if (forheight - j) < 10:
                            forheight = forheight + 1
                    if value is None:
                        checkifemplty = checkifemplty + 1
                        if j < 11 & checkifemplty == 10:
                            checkifempltywidth = checkifempltywidth + 1
                        else:
                            if j > 11 & checkifemplty == 10:
                                break
                    #     if j > excelwidth:
                    #         excelwidth = j
                    # if i > excelheight:
                    #     excelheight = i
            #         print(value)
            #
            # print(sub)
            # print(type(sub))

    finally:
        excel.close()


print("height: " + str(excelheight))
print("width: " + str(excelwidth))
print(str(maxlenght))

tab = [[0 for _ in range(excelwidth+2)] for _ in range((excelheight*2))]
for i in range(maxlenght):
    print("_",end="")
print("")
if file_path:
    print("file:", file_path)
    sstr = file_path.rfind('/')
    sub = file_path[sstr + 1:-1]

    excel = openpyxl.load_workbook(file_path)

    print("what we have:")
    for sheet_name in excel.sheetnames:
        print(sheet_name)

    try:
        name=excel.sheetnames
        for data in name:
            data = excel[data]
            for i in range((excelheight*2)):

                for j in range(excelwidth+2):

                    value = data.cell(row=i + 1, column=j + 1).value
                    if i == 0 | i == excelwidth+1:
                        print("_",end="")
                    else:
                        if value is not None:
                            print(f'value from cell ({i + 1},{j + 1}): {value}')
                    if value is not None:
                        tab[i][j-1]=value

                    #     if j > excelwidth:
                    #         excelwidth = j
                    # if i > excelheight:
                    #     excelheight = i
            #         print(value)
            #
            # print(sub)
            # print(type(sub))

    finally:
        excel.close()

# for i in range(excelheight+2):
#     for j in range(excelwidth+2):
#         if i == 0 | i == excelheight + 1:
#             for i in range(maxlenght*excelwidth):
#                 print("_", end="")
#         else:
#             if j==0:
#                 print("|",end="")
#             else:
#                 if j==excelwidth+1:
#                     print("|")
#                 else:
#
#                     print(str(tab[i-1][j-1]).center(maxlenght), end="")
wfunct=wordfunc.Fun
for i in range(excelheight+2):
    # print(i)
    for j in range(excelwidth+2):
        if i == 0 or i == excelheight + 1:
            print("|",end="")
            for i in range(maxlenght*(excelwidth)):
                print("_", end="")

        else:
            if j==0 and not(i == 0 or i == excelheight + 1):
                print("|",end="")
            else:
                if j==excelwidth+1 and not(i == 0 or i == excelheight + 1):
                    print("|")
                elif (i >=0 and i<=excelheight):
                    # print("i-1: "+str(i)+" j-1: "+str(j))
                    if (wfunct.isfunction(wfunct,tab[i-1][j-2]))==False:
                        print(str(tab[i-1][j-2]).center(maxlenght), end="")

                    else:
                        tabpom=[]
                        command=""
                        alphaCheck=""
                        SubStrOfFun=tab[i-1][j-2]
                        SubStrOfFun=SubStrOfFun[1:]
                        pomvar=0
                        pom2var=0
                        xcel=0
                        pomsum=0
                        ycel=0
                        pom3var=""
                        intlicz=1
                        special_chars = {'+', '-', '=', '/', '*'}
                        # whatfun=wfunct.whatfunction(SubStrOfFun)
                        if wfunct.whatfunction(SubStrOfFun)=="IF":
                            # print("IF")
                            commandstart=SubStrOfFun.rfind("(")
                            subofsub=SubStrOfFun[commandstart:-1]
                            pomvar=commandstart
                            while pom2var==0:

                                if pom2var==1 and len(pom3var)==0:
                                    if subofsub[intlicz]!=" ":
                                        if subofsub[intlicz].isdigit() or (subofsub[intlicz] in special_chars):
                                            ycel=ycel*10+subofsub[intlicz]
                                        else:
                                            if subofsub[intlicz].isalpha():
                                                while subofsub[intlicz].isalpha():
                                                    tabpom.append(subofsub[intlicz])
                                                tabpom2=wfunct.cell(wfunct,tabpom)
                                                for integervalue in tabpom2:
                                                    pomsum=pomsum+tabpom2[integervalue]
                                                xcel=pomsum-1
                                            else:
                                                helpingstrfromtab=tab[ycel,xcel]
                                                tab[ycel,xcel]=helpingstrfromtab[:commandstart+1]+helpingstrfromtab
                                                pom2var=1
                                intlicz=intlicz+1
                                                # alphaCheck=alphaCheck+subofsub[intlicz]

                        elif wfunct.whatfunction(SubStrOfFun)=="OR":
                            print("OR")
                        elif wfunct.whatfunction(SubStrOfFun)=="ROUNDUP":
                            print("ROUND")
                        elif wfunct.whatfunction(SubStrOfFun)=="SUM":
                            print("SUM")
                        else:
                            print(str(tab[i - 1][j - 2]).center(maxlenght), end="")
                    intlicz=intlicz+1

tabbb=[]
tabbb=wfunct.cell(wfunct,"Z2321321ZZZ")
print(tabbb)